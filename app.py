"""
PA-Shift Generator Backend Server
proto-type.py をベースに Flask で API サーバーを構築
複数日対応版
"""

from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from datetime import datetime, timedelta
import copy
import math
import random
import statistics
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

app = Flask(__name__)
CORS(app)  # CORS対応

# スキル3以上を「リーダー」とみなす（スキルは 3=リーダー / 1=アシスタント の2段階運用が前提）
LEADER_SKILL_LEVEL = 3
# desk_max_members / stage_max_members が未設定の場合のデフォルト人数上限（リーダー1名+アシスタント1名）
DEFAULT_DESK_MAX_MEMBERS = 2
DEFAULT_STAGE_MAX_MEMBERS = 2
CONTINUITY_BONUS = 20
# 「進行度（イベント全体を通して今何バンド目まで進んだか）」に対して、本来この時点までに
# 配置されているべき目安回数（ペース目標 = 基準値 × 経過割合）からどれだけ
# 前倒しで進みすぎているか（オーバー分）を計算し、それに応じて優先度を調整する。
# 総量（基準値）だけを見る旧方式だと、序盤に集中して使われても「まだ全体の上限に
# 達していない」うちはペナルティが働かず、少人数の候補が前半に固まって使い切られてしまう
# 問題があったため、「今の時点でのペースを超えていないか」を都度チェックする方式にした。
#
# 前倒し許容幅は「基準値（1人あたりの目安回数）に対する割合」で決める（絶対回数の固定値
# ではない）。固定値にすると、基準値が大きい（例:9回）場合はちょうど良い緩さでも、
# 基準値が小さい場合（例: desk_limit_per_member を5回などに明示設定した場合）には
# 許容幅が基準値の大部分を占めてしまい、前半だけで上限を使い切って後半に誰もいなくなる、
# という問題が起きるため。
# ここまでの前倒し（基準値に対する割合）は特にペナルティを付けない
PACE_SOFT_OVERAGE_RATIO = 0.05
# 割合換算の許容幅が小さくなりすぎないための下限（回数）
PACE_SOFT_OVERAGE_MIN = 0.2
# ソフト超過1につき優先度から差し引く重み（ハード除外はしない・あくまで優先度調整）
PACE_OVERAGE_PENALTY_WEIGHT = 300
# アシスタント（そのロールでリーダーになれないスキルの人）、または明示的な上限設定が
# 適用されているメンバーが、基準値に対してここまで進みすぎている場合は、そのロールの
# 候補からハード除外し、強制的に休ませる（＝中盤・後半に出番を持ち越させる）。
# 自動推定の基準値の場合はリーダーになれるスキルの人には適用しない（リーダー不在による
# バンド不成立を防ぐため、リーダー候補は進みすぎていても優先度を下げるだけに留める）
PACE_HARD_OVERAGE_RATIO = 0.1
# 割合換算の許容幅が小さくなりすぎないための下限（回数）
PACE_HARD_OVERAGE_MIN = 0.35
# 同一役割（卓/ステージ）へ連続で配置してよい上限回数。
# これを超える4回連続以上になる配置は、他に候補がいない場合の最終手段としてのみ許容する
CONSECUTIVE_ROLE_LIMIT = 3
CONSECUTIVE_ROLE_PENALTY = 10000
# 同一役割へまだ上限に達していない範囲で連続させるための優先度ボーナス
# （1バンドごとに役割が頻繁に入れ替わるのを避け、できれば2回以上連続するように後押しする）
ROLE_CONTINUITY_BONUS = 15
# 卓/ステージそれぞれの累計配置回数の差に応じて優先度を補正する重み。
# 「stage_count - desk_count」（卓側）／「desk_count - stage_count」（ステージ側）に掛けることで、
# 片方の役割ばかりに偏っているメンバーには、まだ経験が少ない方の役割の優先度を上げる
# （卓4回・ステージ0回のような極端な役割の型よりを緩和する）
ROLE_BALANCE_WEIGHT = 8
# 直前のバンドで連続して働いていたメンバーが、直前と異なる役割に切り替わるのを抑制するペナルティ。
# 卓→ステージ／ステージ→卓のように、連続勤務中にバンドごとジグザグに役割が入れ替わるのを避けるため、
# 直前と異なる役割の優先度からこの値を引く（直前と同じ役割にはこのペナルティは適用されない）
ROLE_SWITCH_PENALTY = 25

# /api/estimate-limits がおすすめする「卓+ステージ合計配置回数上限」は、
# 卓・ステージそれぞれのリーダー1人あたり目安回数の単純合計に、この倍率をかけて切り上げる。
# 単純合計ちょうど（倍率1.0）だと、NG条件などで一部の人に負担が偏った際にすぐ上限へ
# 抵触してしまい、「リーダー候補が不足している可能性があります」というエラーが起きやすいため、
# 少し余裕を持たせている。
TOTAL_LIMIT_SUGGESTION_MARGIN = 1.5

# 複数パターンを生成し最良案を選ぶ際の試行回数（リクエストの candidate_count で上書き可）
DEFAULT_CANDIDATE_COUNT = 12
MAX_CANDIDATE_COUNT = 50
# 優先度が近い候補者の間で毎回違う組み合わせを試すためのランダムな揺らぎの幅
PRIORITY_JITTER = 4


class _ZeroJitter:
    """揺らぎ無し（従来通りの純粋な貪欲法）の候補を1つは必ず含めるためのダミーRNG"""

    def uniform(self, a, b):
        return 0

    def random(self):
        return 0


_ZERO_JITTER = _ZeroJitter()


def parse_time_slot(time_slot):
    """
    "HH:MM-HH:MM" 形式を datetime の開始/終了に変換
    """
    start_str, end_str = time_slot.split("-")
    start_dt = datetime.strptime(start_str, "%H:%M")
    end_dt = datetime.strptime(end_str, "%H:%M")
    return start_dt, end_dt


def has_time_overlap(slot_a, slot_b):
    """
    2つの時間帯が1分でも重なれば True
    """
    start_a, end_a = parse_time_slot(slot_a)
    start_b, end_b = parse_time_slot(slot_b)
    return start_a < end_b and start_b < end_a


def member_ng_times_for_day(member, day_num=None):
    """
    メンバーのNG時間を日別に取り出す（非破壊）
    """
    ng_times = member.get("ng_times", [])
    if isinstance(ng_times, dict):
        if day_num is None:
            return []
        return ng_times.get(f"day_{day_num}", [])
    if isinstance(ng_times, list):
        return ng_times
    return []


def has_time_conflict(band_slots, member, day_num=None):
    """
    バンド時間帯とメンバーNG時間帯が重複するか
    """
    member_ng_slots = member_ng_times_for_day(member, day_num)
    for band_slot in band_slots:
        for ng_slot in member_ng_slots:
            if has_time_overlap(band_slot, ng_slot):
                return True
    return False


def validate_member_structure(member, index):
    """
    メンバー1件の必須項目と型・値を検証する
    ng_bands / ng_times / req_bands / grade は任意項目（欠けていてもOK）
    """
    required_keys = ["name", "skill_desk", "skill_stage", "count"]
    for key in required_keys:
        if key not in member:
            return f"members[{index}] に必須キー '{key}' がありません"

    if not isinstance(member["name"], str) or not member["name"].strip():
        return f"members[{index}].name は空でない文字列である必要があります"

    numeric_fields = ["skill_desk", "skill_stage", "count"]
    for field in numeric_fields:
        if not isinstance(member[field], (int, float)):
            return f"members[{index}].{field} は数値である必要があります"
        if member[field] < 0:
            return f"members[{index}].{field} は0以上である必要があります"

    ng_bands = member.get("ng_bands", [])
    if ng_bands is not None and not isinstance(ng_bands, list):
        return f"members[{index}].ng_bands は配列である必要があります"

    req_bands = member.get("req_bands", [])
    if req_bands is not None and not isinstance(req_bands, list):
        return f"members[{index}].req_bands は配列である必要があります"

    grade = member.get("grade")
    if grade is not None and not isinstance(grade, (str, int, float)):
        return f"members[{index}].grade は文字列または数値である必要があります"

    ng_times = member.get("ng_times", [])
    if isinstance(ng_times, list):
        for i, slot in enumerate(ng_times):
            if not isinstance(slot, str):
                return f"members[{index}].ng_times[{i}] は文字列である必要があります"
            try:
                parse_time_slot(slot)
            except (ValueError, TypeError):
                return f"members[{index}].ng_times[{i}] は 'HH:MM-HH:MM' 形式である必要があります"
    elif isinstance(ng_times, dict):
        for day_key, slots in ng_times.items():
            if not isinstance(day_key, str) or not day_key.startswith("day_"):
                return f"members[{index}].ng_times のキーは 'day_n' 形式である必要があります"
            if not isinstance(slots, list):
                return f"members[{index}].ng_times['{day_key}'] は配列である必要があります"
            for i, slot in enumerate(slots):
                if not isinstance(slot, str):
                    return f"members[{index}].ng_times['{day_key}'][{i}] は文字列である必要があります"
                try:
                    parse_time_slot(slot)
                except (ValueError, TypeError):
                    return f"members[{index}].ng_times['{day_key}'][{i}] は 'HH:MM-HH:MM' 形式である必要があります"
    else:
        return f"members[{index}].ng_times は配列または日別オブジェクトである必要があります"

    return None


def validate_members(members):
    """
    メンバー配列全体の検証
    """
    if not isinstance(members, list) or len(members) == 0:
        return "メンバーが1人以上必要です"

    for index, member in enumerate(members):
        if not isinstance(member, dict):
            return f"members[{index}] はオブジェクトである必要があります"
        error = validate_member_structure(member, index)
        if error:
            return error
    return None


def normalize_member(member):
    """
    ng_bands / ng_times / req_bands / grade が欠けているメンバーにデフォルト値を補完する（非破壊）
    count / desk_count / stage_count は生成のたびに必ず 0 にリセットする。
    （フロントエンドが前回の生成結果（count が加算済みのmembers）をそのまま次回リクエストに
    使い回すケースがあり、リセットしないと配置回数の上限に誤って抵触し続けてしまうため）
    """
    normalized = dict(member)
    normalized["ng_bands"] = normalized.get("ng_bands") or []
    normalized["ng_times"] = normalized.get("ng_times") or []
    normalized["req_bands"] = normalized.get("req_bands") or []
    normalized.setdefault("grade", None)
    normalized["count"] = 0
    normalized["desk_count"] = 0
    normalized["stage_count"] = 0
    # 直近の配置役割と、その役割への連続配置回数（同一役割への集中を検知するため）
    normalized["last_role"] = None
    normalized["role_streak"] = 0
    normalized["max_role_streak"] = 0
    return normalized


def normalize_members(members):
    """
    メンバー配列全体を正規化する
    """
    return [normalize_member(m) for m in members]


def _validate_optional_non_negative_int(value, field_name):
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, int):
        return f"{field_name} は整数である必要があります"
    if value < 0:
        return f"{field_name} は0以上である必要があります"
    return None


def parse_shift_config(data):
    """
    リクエストボディから任意の詳細設定（人数上限・個人/学年別の配置上限など）を取り出し検証する
    戻り値: (config_dict, error_message)  ※どの項目も省略可能で、省略時は従来通りの動作になる

    卓・ステージの成立条件はスコア閾値ではなく「スキル3以上のリーダーを最低1名含むこと」（ロールベース判定）。
    desk_max_members / stage_max_members は「リーダーを確保した上で何人まで補充するか」の人数上限で、
    未設定の場合は DEFAULT_DESK_MAX_MEMBERS / DEFAULT_STAGE_MAX_MEMBERS を使用する。
    """
    data = data or {}
    config = {
        "desk_max_members": data.get("desk_max_members"),
        "stage_max_members": data.get("stage_max_members"),
        "desk_limit_per_member": data.get("desk_limit_per_member"),
        "stage_limit_per_member": data.get("stage_limit_per_member"),
        "total_limit_per_member": data.get("total_limit_per_member"),
        "grade_limits": data.get("grade_limits", {}) or {},
        "candidate_count": data.get("candidate_count", DEFAULT_CANDIDATE_COUNT),
    }

    for field in (
        "desk_max_members",
        "stage_max_members",
        "desk_limit_per_member",
        "stage_limit_per_member",
        "total_limit_per_member",
    ):
        error = _validate_optional_non_negative_int(config[field], field)
        if error:
            return None, error

    if config["desk_max_members"] is None:
        config["desk_max_members"] = DEFAULT_DESK_MAX_MEMBERS
    if config["stage_max_members"] is None:
        config["stage_max_members"] = DEFAULT_STAGE_MAX_MEMBERS

    if not isinstance(config["grade_limits"], dict):
        return None, "grade_limits はオブジェクトである必要があります"
    for grade_key, limits in config["grade_limits"].items():
        if not isinstance(limits, dict):
            return None, f"grade_limits['{grade_key}'] はオブジェクトである必要があります"
        for role_key in ("desk_max", "stage_max", "total_max"):
            if role_key in limits:
                error = _validate_optional_non_negative_int(
                    limits.get(role_key), f"grade_limits['{grade_key}'].{role_key}"
                )
                if error:
                    return None, error

    candidate_count = config["candidate_count"]
    if isinstance(candidate_count, bool) or not isinstance(candidate_count, int) or candidate_count < 1:
        return None, "candidate_count は1以上の整数である必要があります"
    config["candidate_count"] = max(1, min(candidate_count, MAX_CANDIDATE_COUNT))

    return config, None


def day_sort_key(day_key):
    """
    day_1, day_2, ... を数値としてソートする
    """
    try:
        return int(str(day_key).split("_")[1])
    except (IndexError, ValueError, TypeError):
        return float("inf")


def generate_timetable(start_time_str, band_list, rh_mins, act_mins, break_info=None):
    """
    開始時間とバンドリストから、タイムテーブルを自動生成する関数
    """
    timetable = []
    # 文字列の時間を、計算できる「時計データ」に変換
    current_time = datetime.strptime(start_time_str, "%H:%M")

    for band in band_list:
        # 1. リハの時間を計算して追加
        rh_start = current_time.strftime("%H:%M")
        current_time += timedelta(minutes=rh_mins)  # リハの分数だけ時間を進める
        rh_end = current_time.strftime("%H:%M")
        timetable.append({"time": f"{rh_start}-{rh_end}", "type": "rh", "name": band})

        # 2. 本番の時間を計算して追加
        act_start = current_time.strftime("%H:%M")
        current_time += timedelta(minutes=act_mins)  # 本番の分数だけ時間を進める
        act_end = current_time.strftime("%H:%M")
        timetable.append({"time": f"{act_start}-{act_end}", "type": "act", "name": band})

        # 3. お昼休みの判定（もし指定されていて、今のバンドが終わった直後なら）
        if break_info and break_info.get("after_band") == band:
            break_start = current_time.strftime("%H:%M")
            current_time += timedelta(minutes=break_info["duration"])  # 休憩の分数だけ進める
            break_end = current_time.strftime("%H:%M")
            timetable.append({"time": f"{break_start}-{break_end}", "type": "break", "name": "昼休憩"})

    return timetable


def _band_order_and_times(timetable):
    """
    タイムテーブルから「シフト対象のバンド」と時間帯、前後の順序を整理する
    """
    band_times = {}  # {"バンド名": ["リハ時間", "本番時間"]}
    band_order = []
    for entry in timetable:
        if entry["type"] == "break":
            continue  # 休憩はシフト計算から除外
        b_name = entry["name"]
        if b_name not in band_times:
            band_times[b_name] = []
            band_order.append(b_name)
        band_times[b_name].append(entry["time"])
    return band_order, band_times


def _member_grade_key(member):
    grade = member.get("grade")
    if grade is None:
        return None
    return str(grade)


def _grade_role_limit_ok(member, role, grade_totals, config):
    """
    学年ごとの卓/ステージ合計配置上限（grade_limits）を超えていないか
    """
    grade_key = _member_grade_key(member)
    if grade_key is None:
        return True
    limits = config.get("grade_limits", {}).get(grade_key)
    if not limits:
        return True
    limit = limits.get(f"{role}_max")
    if limit is None:
        return True
    return grade_totals[role][grade_key] < limit


def _grade_total_limit_ok(member, grade_totals, config):
    """
    学年ごとの卓+ステージ合計配置上限（grade_limits[grade].total_max）を超えていないか
    """
    grade_key = _member_grade_key(member)
    if grade_key is None:
        return True
    limits = config.get("grade_limits", {}).get(grade_key)
    if not limits:
        return True
    limit = limits.get("total_max")
    if limit is None:
        return True
    return grade_totals["total"][grade_key] < limit


def _pace_overage(count_so_far, band_index, total_bands, shift_limit):
    """
    「イベント全体を通した進行度（band_index / total_bands）」に対して、
    本来この時点までに配置されているべき目安回数（ペース目標）から
    どれだけ前倒しで進みすぎているかを返す（0以下なら進みすぎていない）。

    band_index は「これまでに処理済みのバンド数」（0始まり・複数日イベントでは
    日をまたいだ通し番号）、total_bands は「イベント全体の総バンド数」。
    """
    if not total_bands or total_bands <= 0 or shift_limit == float("inf"):
        return 0
    progress = band_index / total_bands
    expected_so_far = shift_limit * progress
    return count_so_far - expected_so_far


def _compute_role_pace_baselines(members_data, num_bands, config):
    """
    登録メンバーの「そのロールでリーダーになれる人数／なれない人数（＝アシスタント数）」と
    バンド数から、卓・ステージそれぞれ「リーダー1人あたり」「アシスタント1人あたり」の
    目安配置回数（進行度ベースのペース判定に使う基準値）を計算する。

    卓・ステージとも1バンドにつきリーダー1枠は必須。残り（人数上限-1）枠が補充枠で、
    卓の補充枠は「スキル3未満のアシスタント」専用（_select_team_with_leader の
    assistant_only_fill=True）。ステージの補充枠はスキル不問だが、ここでは
    「本来アシスタントが担うべき想定枠」として同じ考え方で見積もる。

    全メンバー数で一律に割った目安（旧shift_limit）だと、アシスタント適格者が
    少数しかいない場合に実態よりずっと低い目安になり、序盤のうちに「進みすぎ」と
    誤判定されて早期にハード除外されてしまう問題があった。ロール・階層ごとの
    実際の人数で見積もることで、この誤判定を防ぐ。
    """
    desk_max_members = config.get("desk_max_members") or DEFAULT_DESK_MAX_MEMBERS
    stage_max_members = config.get("stage_max_members") or DEFAULT_STAGE_MAX_MEMBERS
    desk_fill_slots = max(0, desk_max_members - 1)
    stage_fill_slots = max(0, stage_max_members - 1)

    def _baseline(skill_field, fill_slots):
        leaders = [m for m in members_data if m.get(skill_field, 0) >= LEADER_SKILL_LEVEL]
        assistants = [m for m in members_data if m.get(skill_field, 0) < LEADER_SKILL_LEVEL]
        leader_avg = (num_bands * 1) / len(leaders) if leaders else None
        assistant_avg = (num_bands * fill_slots) / len(assistants) if assistants else None
        return {
            "leader_count": len(leaders),
            "assistant_count": len(assistants),
            "leader_avg": leader_avg,
            "assistant_avg": assistant_avg,
        }

    return {
        "desk": _baseline("skill_desk", desk_fill_slots),
        "stage": _baseline("skill_stage", stage_fill_slots),
    }


def _pace_target(m, role, skill_field, config, pace_baselines):
    """
    このロール（"desk"/"stage"）のペース判定に使う「1人あたりの目安回数」と、
    どのカウンタ（そのロール単体のカウント か 卓+ステージ合計のカウント）と比較すべきか、
    そして『ユーザーが明示的に設定した上限かどうか』を返す。
    戻り値: (baseline, counter_kind, is_explicit)

    優先順位:
      1. そのロール専用の明示的上限（desk_limit_per_member / stage_limit_per_member）
      2. 合計の明示的上限（total_limit_per_member）
      3. 登録人数・バンド数から自動推定した目安（_compute_role_pace_baselines）
    """
    if role == "desk":
        explicit_role_limit = config.get("desk_limit_per_member")
    else:
        explicit_role_limit = config.get("stage_limit_per_member")
    explicit_total_limit = config.get("total_limit_per_member")

    if explicit_role_limit is not None:
        return explicit_role_limit, "role", True
    if explicit_total_limit is not None:
        return explicit_total_limit, "total", True

    tier = "leader" if m[skill_field] >= LEADER_SKILL_LEVEL else "assistant"
    baseline = pace_baselines.get(role, {}).get(f"{tier}_avg")
    if baseline is None:
        return float("inf"), "role", False
    return baseline, "role", False


def _pace_thresholds(baseline):
    """
    基準値（1人あたりの目安回数）に対する割合で、ソフト/ハード超過の許容幅（回数）を返す。
    基準値が大きくても小さくても、許容幅が「基準値のうちどれだけの割合を前倒しできるか」
    という観点で一定になるようにする（絶対回数の固定値だと、基準値が小さいときに
    相対的に緩くなりすぎ、少ない上限をあっという間に使い切ってしまうため）。
    """
    if baseline == float("inf"):
        return float("inf"), float("inf")
    soft = max(PACE_SOFT_OVERAGE_MIN, baseline * PACE_SOFT_OVERAGE_RATIO)
    hard = max(PACE_HARD_OVERAGE_MIN, baseline * PACE_HARD_OVERAGE_RATIO)
    return soft, hard


def _select_team_with_leader(candidates, skill_field, max_members, assistant_only_fill=False):
    """
    priority降順でソート済みの候補リストから、
    「スキル3以上のリーダーを最低1名含む」チームを組み立てる（ロールベースのハード制約）。

    - リーダー候補（skill_field >= LEADER_SKILL_LEVEL）が1人もいなければ、
      アシスタントが何人いてもチームは組めない（成立不可）ため ([], None) を返す。
    - リーダーが見つかった場合は、そのリーダーを必ず含めた上で、残りの枠を補充する。

    assistant_only_fill=False（既定・ステージ向け）:
      priorityが高い順に、スキルを問わず max_members に達するまで補充する。
    assistant_only_fill=True（卓向け）:
      残りの枠は「スキル3未満のアシスタント」だけで補充する（＝リーダーを2人以上入れない）。
      アシスタント候補がいなければ無理に埋めず、リーダー1名のみのチームも許容する。

    戻り値: (team, leader)  ※team は選ばれたメンバーのリスト、leader は確保できたリーダー（いなければNone）
    """
    if not max_members or max_members < 1:
        return [], None

    leader = next((m for m in candidates if m[skill_field] >= LEADER_SKILL_LEVEL), None)
    if leader is None:
        return [], None

    team = [leader]
    if assistant_only_fill:
        fill_pool = (m for m in candidates if m is not leader and m[skill_field] < LEADER_SKILL_LEVEL)
    else:
        fill_pool = (m for m in candidates if m is not leader)

    for m in fill_pool:
        if len(team) >= max_members:
            break
        team.append(m)

    return team, leader


def _assign_band(
    band, band_slots, prev_band, next_band, prev_assigned, members, grade_totals, config, rng, day_num,
    pace_baselines, band_index, total_bands,
):
    """
    1バンド分の卓・ステージ担当を決定し、メンバーの状態（count/desk_count/stage_count/
    last_role/role_streak/max_role_streak）と学年別合計（grade_totals）を更新する

    band_index / total_bands は「イベント全体を通して今何バンド目まで進んだか」を表し、
    進行度ベースのペース管理（下記）に使う（単日イベントなら total_bands=当日のバンド数、
    複数日イベントなら日をまたいだ通し番号・全日程の総バンド数になる）。
    pace_baselines は `_compute_role_pace_baselines` が返す、登録人数・バンド数から
    自動推定した「卓/ステージ × リーダー/アシスタント」ごとの目安配置回数。

    成立条件（ロールベースのハード制約）:
    卓・ステージそれぞれについて「スキル3以上のリーダーを最低1名含むこと」が必須。
    リーダーを確保できない場合、そのロールは成立不可（アシスタントのみでの編成は不可）。

    卓チームは「リーダー1名 + アシスタント1名」を基本形とし、アシスタント候補がいなければ
    リーダー1名のみの編成も許容する（リーダーを2人重複配置しない）。
    ステージチームは従来通り、リーダー確保後は優先度順にスキルを問わず人数上限まで補充する。

    負担の偏り対策（ソフト制約・優先度調整）:
    - 同一役割に3回連続で入っているメンバーは、さらに同じ役割に入る優先度を大きく下げる
      （shift_limitと同様、ハード除外はせず他に候補がいない場合の最終手段として許容する）
    - 逆に、まだ上限（3回）に達していない範囲では同じ役割を連続させる優先度を少し上げ、
      1バンドごとに役割が頻繁に入れ替わるのを避け、できれば2回以上連続するように後押しする
      （連続を後押しするだけなので、下のロールバランス補正と合わせて「まとまりつつも
      最終的には両役割を経験する」形に寄せる）
    - 卓・ステージそれぞれの優先度に、もう一方の役割の累計配置回数（desk_count/stage_count）
      との差を反映する（ROLE_BALANCE_WEIGHT）。ステージ経験が多く卓経験が少ないメンバーは
      卓の優先度が上がり、逆に卓ばかりのメンバーは卓の優先度が下がる。これにより
      「卓ばかり／ステージばかり」という極端な役割の偏りが緩和される
    - 直前のバンドで連続して働いていたメンバーが、直前と異なる役割に切り替わる優先度を
      下げる（ROLE_SWITCH_PENALTY）。これにより、連続勤務中に卓→ステージ→卓...のように
      バンドごとジグザグに役割が入れ替わるのを避け、連続勤務中はなるべく同じ役割にとどまる
      ようにする（役割継続ボーナスと組み合わさることで効果が強まる）

    配置回数のペース判定（_pace_target / _pace_overage）は「今の進行度（band_index /
    total_bands）」に対して、そのロール専用の明示的上限（desk/stage_limit_per_member）→
    合計の明示的上限（total_limit_per_member）→ 登録人数・バンド数から自動推定した
    ロール・階層別の目安（_compute_role_pace_baselines）の優先順位で基準値を決め、
    ペースを超えて前倒しで働いているメンバーには優先度ペナルティが付く。さらに、
    「そのロールでリーダーになれないスキル（＝アシスタント）」のメンバーが大幅に超過して
    いる場合は、そのロールの候補からハード除外して強制的に休ませる（中盤・後半に出番を
    持ち越す）。基準値がユーザーの明示的な上限設定に基づく場合はリーダーにもハード除外を
    適用するが、自動推定の目安に基づく場合はリーダーには適用せず優先度を下げるだけに
    留める（リーダー不在によるバンド不成立を防ぐため）。
    """
    desk_max_members = config.get("desk_max_members", DEFAULT_DESK_MAX_MEMBERS)
    stage_max_members = config.get("stage_max_members", DEFAULT_STAGE_MAX_MEMBERS)
    desk_limit_per_member = config.get("desk_limit_per_member")
    stage_limit_per_member = config.get("stage_limit_per_member")
    total_limit_per_member = config.get("total_limit_per_member")

    def base_eligible(m):
        # NG判定①：自分が出演するバンド、またはその「前後」ならシフト不可
        if band in m["ng_bands"] or prev_band in m["ng_bands"] or next_band in m["ng_bands"]:
            return False
        # NG判定②：NG時間に1分でも被っているか（部分一致判定）
        if has_time_conflict(band_slots, m, day_num):
            return False
        return True

    def _pace_count_so_far(m, role, counter_kind):
        if counter_kind == "total":
            return m["count"]
        return m["desk_count"] if role == "desk" else m["stage_count"]

    def pace_penalty(m, role, skill_field):
        # 「今何バンド目まで進んだか」に対する目安ペースからどれだけ前倒しで進みすぎて
        # いるかに応じて優先度を下げる（総量ではなく進行度ベース）。ハード除外はしない
        # ＝他に候補がいなければ、進みすぎていてもこの人が選ばれる余地は残す。
        baseline, counter_kind, _ = _pace_target(m, role, skill_field, config, pace_baselines)
        soft_overage, _ = _pace_thresholds(baseline)
        overage = _pace_overage(_pace_count_so_far(m, role, counter_kind), band_index, total_bands, baseline)
        if overage <= soft_overage:
            return 0
        return (overage - soft_overage) * PACE_OVERAGE_PENALTY_WEIGHT

    def pace_hard_exclude(m, role, skill_field):
        # アシスタント（そのロールでリーダーになれないスキル）がペースを大幅に超えて
        # 進みすぎている場合は、そのロールの候補から完全に除外し、強制的に休ませる。
        # 基準値が自動推定（ユーザー未設定）の場合はリーダーには適用しない
        # （リーダー不在によるバンド不成立を防ぐ）。ユーザーが明示的に上限を設定した
        # 場合は、その上限を前倒しで使い切らないようリーダーにも適用する。
        baseline, counter_kind, is_explicit = _pace_target(m, role, skill_field, config, pace_baselines)
        if not is_explicit and m[skill_field] >= LEADER_SKILL_LEVEL:
            return False
        _, hard_overage = _pace_thresholds(baseline)
        overage = _pace_overage(_pace_count_so_far(m, role, counter_kind), band_index, total_bands, baseline)
        return overage > hard_overage

    def consecutive_role_penalty(m, role):
        # 同一役割に既に3回連続で入っているメンバーが、さらに同じ役割に入るのを強く抑制する
        # （ハード除外はせず、他に候補がいない場合の最終手段としてのみ許容する）
        if m.get("last_role") == role and m.get("role_streak", 0) >= CONSECUTIVE_ROLE_LIMIT:
            return CONSECUTIVE_ROLE_PENALTY
        return 0

    def role_continuity_bonus(m, role):
        # 直前と同じ役割で、かつまだ連続上限（3回）に達していなければ、
        # その役割を継続する優先度を少し上げる（1バンドごとの頻繁な入れ替わりを避けるため）
        if m.get("last_role") == role and 0 < m.get("role_streak", 0) < CONSECUTIVE_ROLE_LIMIT:
            return ROLE_CONTINUITY_BONUS
        return 0

    def role_switch_penalty(m, role):
        # 直前のバンドで連続して働いており（prev_assignedに含まれる）、かつ直前の役割が
        # 今回スコアリングしている役割と異なる場合にペナルティを課す。
        # 連続勤務のたびに卓→ステージ→卓...とジグザグに入れ替わるのを避け、
        # 一度どちらかの役割に入ったら、連続勤務中はその役割にとどまりやすくする
        if m["name"] in prev_assigned and m.get("last_role") is not None and m.get("last_role") != role:
            return ROLE_SWITCH_PENALTY
        return 0

    # 卓チーム編成
    available_desk = []
    for m in members:
        if not base_eligible(m):
            continue
        if desk_limit_per_member is not None and m["desk_count"] >= desk_limit_per_member:
            continue
        if total_limit_per_member is not None and m["count"] >= total_limit_per_member:
            continue
        if not _grade_role_limit_ok(m, "desk", grade_totals, config):
            continue
        if not _grade_total_limit_ok(m, grade_totals, config):
            continue
        if pace_hard_exclude(m, "desk", "skill_desk"):
            continue

        # 「これまで卓に何回入ったか」だけを見て優先度を下げる（卓+ステージの合計回数では
        # なく卓回数そのもの）。合計回数で下げてしまうと、卓のアシスタント予約枠で頻繁に
        # 卓へ配置される初心者ほど合計回数が早く積み上がり、結果としてステージの優先度まで
        # 一緒に下がってしまい、いつまで経ってもステージに回れなくなる（下のロールバランス
        # 補正と効果が矛盾してしまう）ため、卓は卓回数、ステージはステージ回数で判定する。
        priority_desk = -m["desk_count"] * 10
        if band in m["req_bands"]:
            priority_desk += 100
        priority_desk += m["skill_desk"]
        if m["name"] in prev_assigned:
            priority_desk += CONTINUITY_BONUS
        priority_desk -= pace_penalty(m, "desk", "skill_desk")
        priority_desk -= consecutive_role_penalty(m, "desk")
        priority_desk += role_continuity_bonus(m, "desk")
        priority_desk -= role_switch_penalty(m, "desk")
        # ステージ経験が卓経験より多いほど卓の優先度を上げ、逆に卓ばかりなら下げる
        priority_desk += (m["stage_count"] - m["desk_count"]) * ROLE_BALANCE_WEIGHT
        priority_desk += rng.uniform(0, PRIORITY_JITTER)

        candidate = m.copy()
        candidate["priority_desk"] = priority_desk
        available_desk.append(candidate)

    available_desk.sort(key=lambda x: x["priority_desk"], reverse=True)
    desk_team, desk_leader = _select_team_with_leader(
        available_desk, "skill_desk", desk_max_members, assistant_only_fill=True
    )

    # ステージチーム編成（すでに卓に割り当てられたメンバーは除く）
    desk_team_names = {m["name"] for m in desk_team}
    available_stage = []
    for m in members:
        if m["name"] in desk_team_names:
            continue
        if not base_eligible(m):
            continue
        if stage_limit_per_member is not None and m["stage_count"] >= stage_limit_per_member:
            continue
        if total_limit_per_member is not None and m["count"] >= total_limit_per_member:
            continue
        if not _grade_role_limit_ok(m, "stage", grade_totals, config):
            continue
        if not _grade_total_limit_ok(m, grade_totals, config):
            continue
        if pace_hard_exclude(m, "stage", "skill_stage"):
            continue

        # 卓と同様に、合計回数ではなく「これまでステージに何回入ったか」だけで優先度を下げる。
        # 卓のアシスタント予約枠で卓回数（＝合計回数）が多くなりがちな初心者が、合計回数基準の
        # 減点によってステージの優先度まで下げられてしまうと、ステージにはいつまでも回って
        # こなくなる（下のロールバランス補正で押し上げても、この減点で相殺されてしまう）。
        priority_stage = -m["stage_count"] * 10
        if band in m["req_bands"]:
            priority_stage += 100
        # 卓と違い、ここに priority_desk のような「+m['skill_stage']」（スキル値そのままの加点）は
        # 入れない。ステージはリーダー確保後の残り枠を「スキルを問わず優先度順」で補充する方針
        # のため、もしスキル値を直接加点すると、スキル3以上の人が常にスキル1〜2の人より
        # 2〜4点も有利になってしまい、リーダー確保後の枠がほぼ常に別のリーダー候補で
        # 埋まってしまう（初心者がステージにほとんど配置されない）という偏りの原因になっていた。
        # スキル差はリーダー成立条件（最低1名は必須）だけで担保し、残り枠は継続性・役割バランス・
        # ペース・公平性などスキル以外の要素だけで、スキル3以上の人と初心者を同列に競わせる。
        if m["name"] in prev_assigned:
            priority_stage += CONTINUITY_BONUS
        priority_stage -= pace_penalty(m, "stage", "skill_stage")
        priority_stage -= consecutive_role_penalty(m, "stage")
        priority_stage += role_continuity_bonus(m, "stage")
        priority_stage -= role_switch_penalty(m, "stage")
        # 卓経験がステージ経験より多いほどステージの優先度を上げ、逆にステージばかりなら下げる
        priority_stage += (m["desk_count"] - m["stage_count"]) * ROLE_BALANCE_WEIGHT
        priority_stage += rng.uniform(0, PRIORITY_JITTER)

        candidate = m.copy()
        candidate["priority_stage"] = priority_stage
        available_stage.append(candidate)

    available_stage.sort(key=lambda x: x["priority_stage"], reverse=True)
    stage_team, stage_leader = _select_team_with_leader(available_stage, "skill_stage", stage_max_members)

    # 状態更新（カウント・学年別合計・連続役割ストリーク）
    desk_names = [m["name"] for m in desk_team]
    stage_names = [m["name"] for m in stage_team]
    assigned_names = set(desk_names) | set(stage_names)
    for m in members:
        if m["name"] in assigned_names:
            m["count"] += 1
            grade_key = _member_grade_key(m)
            if grade_key is not None:
                grade_totals["total"][grade_key] += 1
        if m["name"] in desk_names:
            m["desk_count"] += 1
            grade_key = _member_grade_key(m)
            if grade_key is not None:
                grade_totals["desk"][grade_key] += 1
        if m["name"] in stage_names:
            m["stage_count"] += 1
            grade_key = _member_grade_key(m)
            if grade_key is not None:
                grade_totals["stage"][grade_key] += 1

        # 配置されなかったバンドではストリークを維持し、
        # 「連続で働いたときに同じ役割が何回続いたか」を追跡する
        if m["name"] in desk_names:
            assigned_role = "desk"
        elif m["name"] in stage_names:
            assigned_role = "stage"
        else:
            assigned_role = None
        if assigned_role is not None:
            if m.get("last_role") == assigned_role:
                m["role_streak"] = m.get("role_streak", 0) + 1
            else:
                m["role_streak"] = 1
            m["last_role"] = assigned_role
            if m["role_streak"] > m.get("max_role_streak", 0):
                m["max_role_streak"] = m["role_streak"]

    band_result = {"卓": desk_names, "ステージ": stage_names}

    infeasible_entry = None
    if desk_leader is None or stage_leader is None:
        reasons = []
        if desk_leader is None:
            reasons.append(
                f"卓にスキル{LEADER_SKILL_LEVEL}以上のリーダーを配置できませんでした"
                "（NG条件・学年/個人ごとの配置上限などにより、対応できるリーダー候補が不足している可能性があります）"
            )
        if stage_leader is None:
            reasons.append(
                f"ステージにスキル{LEADER_SKILL_LEVEL}以上のリーダーを配置できませんでした"
                "（NG条件・学年/個人ごとの配置上限などにより、対応できるリーダー候補が不足している可能性があります）"
            )
        infeasible_entry = {
            "band": band,
            "desk_has_leader": desk_leader is not None,
            "stage_has_leader": stage_leader is not None,
            "desk_members": desk_names,
            "stage_members": stage_names,
            "reason": " / ".join(reasons),
        }

    team_size = len(desk_names) + len(stage_names)
    return band_result, team_size, infeasible_entry


def _run_day_core(timetable, members, grade_totals, config, rng, pace_baselines, day_num=None, band_index_offset=0, total_bands=None):
    """
    1日分のタイムテーブルに対して、バンド順にシフトを組み立てる（コア処理・非候補探索）

    band_index_offset / total_bands は「イベント全体を通した進行度」を計算するために使う。
    単日イベントでは band_index_offset=0・total_bands=当日のバンド数（省略時は自動算出）、
    複数日イベントでは呼び出し側が日をまたいだ通し番号と全日程の総バンド数を渡す。
    pace_baselines は `_compute_role_pace_baselines` の戻り値。
    """
    band_order, band_times = _band_order_and_times(timetable)
    if total_bands is None:
        total_bands = len(band_order)

    day_shift = {}
    infeasible_bands = []
    band_team_sizes = []

    for i, band in enumerate(band_order):
        prev_band = band_order[i - 1] if i > 0 else None
        next_band = band_order[i + 1] if i < len(band_order) - 1 else None

        prev_assigned = set()
        if prev_band and prev_band in day_shift:
            prev_assigned = set(day_shift[prev_band]["卓"] + day_shift[prev_band]["ステージ"])

        band_result, team_size, infeasible_entry = _assign_band(
            band, band_times[band], prev_band, next_band, prev_assigned,
            members, grade_totals, config, rng, day_num, pace_baselines,
            band_index_offset + i, total_bands,
        )
        day_shift[band] = band_result
        band_team_sizes.append(team_size)
        if infeasible_entry:
            infeasible_bands.append(infeasible_entry)

    return day_shift, infeasible_bands, band_team_sizes


def _candidate_quality(infeasible_count, band_team_sizes, members):
    """
    候補案の良さを比較するためのスコア（小さいほど良い）
    1) 成立しなかったバンド数が少ない方が良い
    2) 同一役割への連続配置が3回を超えているメンバーが少ない・超過幅が小さい方が良い
       （卓・ステージのどちらかが0回のまま偏るのは許容するため、ここでは評価しない）
    3) バンドごとの人数がバンド間で均されている方が良い（標準偏差が小さい）
    4) メンバー間の配置回数の偏りが少ない方が良い（標準偏差が小さい）
    """
    consecutive_overrun = sum(
        max(0, m.get("max_role_streak", 0) - CONSECUTIVE_ROLE_LIMIT) for m in members
    )
    balance = statistics.pstdev(band_team_sizes) if len(band_team_sizes) > 1 else 0.0
    counts = [m["count"] for m in members]
    fairness = statistics.pstdev(counts) if len(counts) > 1 else 0.0
    return (infeasible_count, consecutive_overrun, balance, fairness)


def generate_pa_shift(timetable, members_data, day_num=None, config=None):
    """
    PAシフトを作成する関数（v4）
    - NG時間は部分一致（1分でも被れば除外）で判定
    - 卓・ステージの成立条件はロールベース：スキル3以上のリーダーを最低1名含むことが必須
    - 同一役割（卓/ステージ）への連続配置は3回までに抑えつつ、上限に達するまではできれば
      2回以上連続するよう後押しする（卓・ステージのどちらかが0回のまま偏るのはOK）
    - 配置回数は「イベント全体の進行度に対するペース」で管理する。ペース基準値は
      登録人数・バンド数から卓/ステージ×リーダー/アシスタント別に自動推定し
      （_compute_role_pace_baselines）、ペースを超えて前倒しで働いているアシスタント
      （そのロールでリーダーになれないスキルの人）は、大幅に超過すると一時的に候補から
      除外して休ませる（特定の人が序盤に集中するのを防ぐ）
    - 優先度にランダムな揺らぎを加えた複数パターンを生成し、
      「成立バンド数が多い」→「連続配置の超過が少ない」→
      「バンド間の人数の均一性が高い」→「配置回数の公平性が高い」
      の順で最も良い案を採用する
    """
    if config is None:
        config = {"desk_max_members": DEFAULT_DESK_MAX_MEMBERS, "stage_max_members": DEFAULT_STAGE_MAX_MEMBERS}

    band_order, _ = _band_order_and_times(timetable)
    num_bands = len(band_order)
    pace_baselines = _compute_role_pace_baselines(members_data, num_bands, config)

    rng = random.Random()
    candidate_count = config.get("candidate_count", DEFAULT_CANDIDATE_COUNT)

    best = None
    for i in range(candidate_count):
        members_state = normalize_members(copy.deepcopy(members_data))
        grade_totals = {"desk": defaultdict(int), "stage": defaultdict(int), "total": defaultdict(int)}
        candidate_rng = _ZERO_JITTER if i == 0 else rng

        day_shift, infeasible_bands, band_team_sizes = _run_day_core(
            timetable, members_state, grade_totals, config, candidate_rng, pace_baselines, day_num,
            band_index_offset=0, total_bands=num_bands,
        )
        quality = _candidate_quality(len(infeasible_bands), band_team_sizes, members_state)
        if best is None or quality < best[0]:
            best = (quality, day_shift, members_state, infeasible_bands)

    _, day_shift, members_state, infeasible_bands = best
    return day_shift, members_state, infeasible_bands


def generate_timetable_multi_day(timetable_config):
    """
    複数日のタイムテーブルを生成する関数
    timetable_config: {
        "num_days": 2,
        "days": [
            {
                "day_number": 1,
                "start_time": "11:30",
                "bands": ["band1", "band2"],
                "rh_mins": 15,
                "act_mins": 10,
                "break_duration": 60,
                "break_after_band": "band1"
            },
            {...}
        ]
    }
    """
    result = {}
    for day_config in timetable_config.get("days", []):
        day_num = day_config.get("day_number", 1)
        timetable = generate_timetable(
            day_config.get("start_time", "11:30"),
            day_config.get("bands", []),
            day_config.get("rh_mins", 15),
            day_config.get("act_mins", 10),
            {
                "after_band": day_config.get("break_after_band", ""),
                "duration": day_config.get("break_duration", 60)
            } if day_config.get("break_after_band") else None
        )
        result[f"day_{day_num}"] = timetable
    return result


def generate_pa_shift_multi_day(timetable_multi, members_data, config=None):
    """
    複数日のシフトを生成する関数（v4）
    - ng_times は日別dictのまま非破壊で参照する（2日目以降もNG時間が正しく効く）
    - 卓・ステージの成立条件はロールベース：スキル3以上のリーダーを最低1名含むことが必須
    - 配置回数の上限（shift_limit）・個人/学年別の配置上限は「全日程を通じた合計」で管理する
    - 同一役割（卓/ステージ）への連続配置は日をまたいでも3回までに抑えつつ、上限に達するまでは
      できれば2回以上連続するよう後押しする（卓・ステージのどちらかが0回のまま偏るのはOK）
    - 配置回数は「全日程を通した進行度に対するペース」で管理する（日をまたいだ通し番号で
      判定するため、特定の1日だけに特定のメンバーが集中する偏りも抑制される）。ペース基準値は
      登録人数・全日程の総バンド数から卓/ステージ×リーダー/アシスタント別に自動推定し
      （_compute_role_pace_baselines）、ペースを超えて前倒しで働いているアシスタントは、
      大幅に超過すると一時的に候補から除外して休ませる
    - 優先度にランダムな揺らぎを加えた複数パターンを全日程分通しで生成し、最も良い案を採用する
    """
    if config is None:
        config = {"desk_max_members": DEFAULT_DESK_MAX_MEMBERS, "stage_max_members": DEFAULT_STAGE_MAX_MEMBERS}

    sorted_days = sorted(timetable_multi.items(), key=lambda item: day_sort_key(item[0]))

    total_bands = sum(len(_band_order_and_times(timetable)[0]) for _, timetable in sorted_days)
    pace_baselines = _compute_role_pace_baselines(members_data, total_bands, config)

    rng = random.Random()
    candidate_count = config.get("candidate_count", DEFAULT_CANDIDATE_COUNT)

    best = None
    for i in range(candidate_count):
        members_state = normalize_members(copy.deepcopy(members_data))
        grade_totals = {"desk": defaultdict(int), "stage": defaultdict(int), "total": defaultdict(int)}
        candidate_rng = _ZERO_JITTER if i == 0 else rng

        shift_result = {}
        infeasible_days = {}
        all_band_team_sizes = []
        band_index_offset = 0

        for day_key, timetable in sorted_days:
            day_num = day_sort_key(day_key)
            day_shift, infeasible_bands, band_team_sizes = _run_day_core(
                timetable, members_state, grade_totals, config, candidate_rng, pace_baselines, day_num,
                band_index_offset=band_index_offset, total_bands=total_bands,
            )
            band_index_offset += len(band_team_sizes)
            shift_result[day_key] = day_shift
            all_band_team_sizes.extend(band_team_sizes)
            if infeasible_bands:
                infeasible_days[day_key] = infeasible_bands

        total_infeasible = sum(len(v) for v in infeasible_days.values())
        quality = _candidate_quality(total_infeasible, all_band_team_sizes, members_state)
        if best is None or quality < best[0]:
            best = (quality, shift_result, members_state, infeasible_days)

    _, shift_result, members_state, infeasible_days = best
    return shift_result, members_state, infeasible_days


def create_excel_workbook(timetable_multi, shift_result, members):
    """
    複数日のシフトをエクセルワークブックとして作成
    各日ごとに1シート、最後に集計シートを追加
    """
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # スタイル定義
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 各日のシートを作成
    day_sheets = {}
    for day_key, timetable in sorted(timetable_multi.items(), key=lambda item: day_sort_key(item[0])):
        day_num = int(day_key.split('_')[1])
        sheet_name = f"{day_num}日目"
        ws = wb.create_sheet(sheet_name)
        day_sheets[day_key] = ws

        # ヘッダー行
        headers = ["時間帯", "種別", "バンド名", "担当（卓）", "担当（ステージ）"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # データ行
        for row, entry in enumerate(timetable, 2):
            assigned = shift_result[day_key].get(entry["name"], {"卓": [], "ステージ": []})
            desk = "、".join(assigned["卓"]) if assigned["卓"] else "-"
            stage = "、".join(assigned["ステージ"]) if assigned["ステージ"] else "-"
            entry_type = "リハ" if entry["type"] == "rh" else "本番" if entry["type"] == "act" else "休憩"

            row_data = [entry["time"], entry_type, entry["name"], desk, stage]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 列幅を調整
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

    # 全体集計シートを作成
    ws_summary = wb.create_sheet("全体集計", 0)

    # 集計ヘッダー
    summary_headers = ["メンバー名", "学年", "シフト回数", "卓回数", "ステージ回数"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # メンバーをカウント値で降順ソート
    sorted_members = sorted(members, key=lambda x: x["count"], reverse=True)

    # データ行
    for row, member in enumerate(sorted_members, 2):
        member_data = [
            member["name"],
            member.get("grade") if member.get("grade") is not None else "-",
            member["count"],
            member.get("desk_count", "-"),
            member.get("stage_count", "-"),
        ]
        for col, value in enumerate(member_data, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 14

    return wb




@app.route("/", methods=["GET"])
def index():
    """
    ルートURL: templates/index.html（フロントエンド）を返す
    """
    return render_template("index.html")


@app.route("/api/generate-shift-multi-day", methods=["POST"])
def api_generate_shift_multi_day():
    """
    複数日のタイムテーブルとシフトを生成する統合エンドポイント
    """
    try:
        data = request.json

        # リクエストデータの取得
        num_days = data.get("num_days", 1)
        days = data.get("days", [])
        members = data.get("members", [])

        # バリデーション
        if num_days < 1:
            return jsonify({"error": "イベント日数は1日以上である必要があります"}), 400
        if not days or len(days) == 0:
            return jsonify({"error": "各日のバンド設定が必要です"}), 400
        members_error = validate_members(members)
        if members_error:
            return jsonify({"error": members_error}), 400

        # 詳細設定（人数上限・個人/学年別の配置上限など、すべて任意）の取得
        config, config_error = parse_shift_config(data)
        if config_error:
            return jsonify({"error": config_error}), 400

        # 複数日タイムテーブル生成
        timetable_multi = generate_timetable_multi_day({"num_days": num_days, "days": days})

        # 複数日シフト生成
        shift_result, updated_members, infeasible_days = generate_pa_shift_multi_day(timetable_multi, members, config)
        if infeasible_days:
            return (
                jsonify(
                    {
                        "error": "条件を満たすシフトを生成できませんでした（成立しなかったバンドがあります。詳細は infeasible_days を確認してください）",
                        "infeasible_days": infeasible_days,
                        "timetable_multi": timetable_multi,
                        "shift": shift_result,
                        "members": updated_members,
                    }
                ),
                400,
            )

        # レスポンス作成
        return jsonify(
            {
                "status": "success",
                "timetable_multi": timetable_multi,
                "shift": shift_result,
                "members": updated_members,
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download-excel", methods=["POST"])
def api_download_excel():
    """
    複数日のシフトをエクセルファイルとしてダウンロード
    """
    try:
        data = request.json

        # リクエストデータの取得
        timetable_multi = data.get("timetable_multi", {})
        shift = data.get("shift", {})
        members = data.get("members", [])

        # バリデーション
        if not timetable_multi or not shift:
            return jsonify({"error": "シフトデータが必要です"}), 400

        # エクセルワークブック作成
        wb = create_excel_workbook(timetable_multi, shift, members)

        # BytesIOに出力
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ファイルとして返す
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"pa_shift_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-shift", methods=["POST"])
def api_generate_shift():
    """
    タイムテーブル生成とシフト生成を行う統合エンドポイント
    """
    try:
        data = request.json

        # リクエストデータの取得
        start_time = data.get("start_time", "11:30")
        bands = data.get("bands", [])
        rh_mins = data.get("rh_mins", 15)
        act_mins = data.get("act_mins", 10)
        break_duration = data.get("break_duration", 60)
        break_after_band = data.get("break_after_band", "")
        members = data.get("members", [])

        # バリデーション
        if not bands:
            return jsonify({"error": "バンドが1つ以上必要です"}), 400
        members_error = validate_members(members)
        if members_error:
            return jsonify({"error": members_error}), 400

        # 詳細設定（人数上限・個人/学年別の配置上限など、すべて任意）の取得
        config, config_error = parse_shift_config(data)
        if config_error:
            return jsonify({"error": config_error}), 400

        # breakInfoの組み立て
        break_info = None
        if break_after_band:
            break_info = {"after_band": break_after_band, "duration": break_duration}

        # タイムテーブル生成
        timetable = generate_timetable(start_time, bands, rh_mins, act_mins, break_info)

        # シフト生成
        shift_result, updated_members, infeasible_bands = generate_pa_shift(timetable, members, config=config)
        if infeasible_bands:
            return (
                jsonify(
                    {
                        "error": "条件を満たすシフトを生成できませんでした（成立しなかったバンドがあります。詳細は infeasible_bands を確認してください）",
                        "infeasible_bands": infeasible_bands,
                        "timetable": timetable,
                        "shift": shift_result,
                        "members": updated_members,
                    }
                ),
                400,
            )

        # レスポンス作成
        return jsonify(
            {
                "status": "success",
                "timetable": timetable,
                "shift": shift_result,
                "members": updated_members,
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/estimate-limits", methods=["POST"])
def api_estimate_limits():
    """
    シフトは生成せず、登録済みのバンド・メンバー情報から
    「卓/ステージ × リーダー/アシスタント」1人あたりの目安配置回数と、
    詳細設定（合計回数などの上限）に使えるおすすめのデフォルト値を計算して返す軽量エンドポイント。

    合計回数のおすすめ値は、登録人数のうちスキル3以上（リーダー）の人数を基準に
    「1人あたり平均で何回必要か」を計算したもの（多くの場合リーダーが大半を占めるため、
    これを既定値にするのが実用的）。アシスタントの目安は参考値として別途返す。
    """
    try:
        data = request.json or {}
        members = data.get("members", [])
        if not members:
            return jsonify({"error": "メンバーが1人以上必要です"}), 400
        members_error = validate_members(members)
        if members_error:
            return jsonify({"error": members_error}), 400

        # 単日（bands）・複数日（days）どちらの指定でも受け付ける
        if "days" in data:
            num_bands = sum(len(day.get("bands", [])) for day in data.get("days", []))
        else:
            num_bands = len(data.get("bands", []))
        if num_bands <= 0:
            return jsonify({"error": "バンドが1つ以上必要です"}), 400

        config, config_error = parse_shift_config(data)
        if config_error:
            return jsonify({"error": config_error}), 400

        baselines = _compute_role_pace_baselines(members, num_bands, config)

        def _round1(value):
            return round(value, 1) if value is not None else None

        def _suggest_ceiling(value):
            # 「平均回数ちょうど」だと端数分だけ全員が確実に上限に届かなくなるため、
            # 少し余裕を持たせて切り上げた値をおすすめのデフォルトにする
            if value is None:
                return None
            return math.ceil(value)

        desk = baselines["desk"]
        stage = baselines["stage"]

        response = {
            "num_bands": num_bands,
            "desk": {
                "leader_count": desk["leader_count"],
                "assistant_count": desk["assistant_count"],
                "leader_avg": _round1(desk["leader_avg"]),
                "assistant_avg": _round1(desk["assistant_avg"]),
            },
            "stage": {
                "leader_count": stage["leader_count"],
                "assistant_count": stage["assistant_count"],
                "leader_avg": _round1(stage["leader_avg"]),
                "assistant_avg": _round1(stage["assistant_avg"]),
            },
            "suggested_desk_limit_per_member": _suggest_ceiling(desk["leader_avg"]),
            "suggested_stage_limit_per_member": _suggest_ceiling(stage["leader_avg"]),
            "suggested_total_limit_per_member": _suggest_ceiling(
                ((desk["leader_avg"] or 0) + (stage["leader_avg"] or 0)) * TOTAL_LIMIT_SUGGESTION_MARGIN
            ),
        }
        return jsonify(response)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/health", methods=["GET"])
def health_check():
    """ヘルスチェック"""
    return jsonify({"status": "ok", "message": "PA-Shift Backend is running"})


if __name__ == "__main__":
    print("🚀 PA-Shift Backend Server starting...")
    print("📌 http://localhost:5000 でサーバーが起動しました")
    print("🔗 http://localhost:5000/api/health でヘルスチェック")
    app.run(debug=True, port=5000)
